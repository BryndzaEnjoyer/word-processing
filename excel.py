import openpyxl as xl
from config import get_data_file_path
import pandas as pd
path=get_data_file_path('Výpočty a filtrovanie.xlsx')

wb = xl.load_workbook(path)
print(wb.sheetnames)


def select_sheet(x):
    wb._active_sheet_index = x
    print("Active sheet: ", wb.active)
    sheet_obj = wb.active
    row = sheet_obj.max_row
    column = sheet_obj.max_column


    all_values = []
    for i in range(1, row + 1):
        row_values = []
        for j in range(1, column + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            row_values.append(cell_obj.value)
        all_values.append(row_values)

    return all_values

y=int(input("select a sheet number:"))

data=select_sheet(y)



df = pd.DataFrame(data)
df.drop(9, axis=1, inplace=True)
df.drop(10, axis=1, inplace=True)
df.drop(8, axis=1, inplace=True)
df.drop(7, axis=1, inplace=True)
for i in range(7,16):
    df.drop(i, axis=0, inplace=True)
df.drop(0, axis=0, inplace=True)

print(df)




