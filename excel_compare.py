from openpyxl import Workbook as wb
from openpyxl import load_workbook
from pathlib import Path
from config import get_data_file_path, get_output_file_path





def compare_results(filename_data, filename_output):
    doc_data = load_workbook(get_data_file_path(filename_data), data_only=False)
    doc_output = load_workbook(get_output_file_path(filename_output), data_only=False)
    sheets_a, sheets_b = doc_data.sheetnames, doc_output.sheetnames
    common = [s for s in sheets_a if s in sheets_b]
    conflicts = {}
    for s in common:
        sheet_data = doc_data[s]
        sheet_output = doc_output[s]
        for row in sheet_data.iter_rows():
            for cell in row:
                if cell.value != sheet_output[cell.coordinate].value:
                    conflicts[cell.coordinate] = (cell.value, sheet_output[cell.coordinate].value)

    return conflicts

if __name__ == "__main__":
    print (compare_results("Výpočty a filtrovanie.xlsx", "Výpočty a filtrovanie_RIEŠENIE.xlsx"))